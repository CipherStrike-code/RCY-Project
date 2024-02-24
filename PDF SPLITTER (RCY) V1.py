import os
import PyPDF2
import openpyxl

# Input parameters
try:
    while True:
        # Enter the PDF file required and remove the quotation at the start and end of the input
        selected_pdf_file = input(
            "Enter the directory of the pdf you want to split: "
        ).strip('"')
        if selected_pdf_file.endswith(".pdf"):
            if selected_pdf_file == "":
                print("Filename cannot be empty.")
                continue

            if not os.path.exists(selected_pdf_file):
                print("Selected PDF does not exist.")
                continue

            # If the code reaches here, it means the input is valid
            break
        else:
            print("Please enter a valid pdf file with the .pdf extension.")
            continue

    pdf_filter = "*.pdf"

    while True:
        # Enter the excel file required and remove the quotation at the start and end of the input
        excel_path = input(
            "Enter the directory of the excel file you would like to use: "
        ).strip('"')

        # Checking if the input is empty
        if not excel_path:
            print("Filename cannot be empty.")
            continue

        # Check if the file exists
        if not os.path.exists(excel_path):
            print("Selected Excel file does not exist.")
            continue

        break

    # Allow the user to choose a specific sheet
    excel_instance = openpyxl.load_workbook(excel_path)
    available_sheets = excel_instance.sheetnames

    print("Available sheets in the Excel file:")
    for sheet in available_sheets:
        print(sheet)

    while True:
        selected_sheet = input("Enter the name of the sheet you want to use: ")

        if selected_sheet in available_sheets:
            break
        else:
            print("Invalid sheet name. Please enter a valid sheet name.")
            continue

    # Get the selected sheet
    particulars_sheet = excel_instance[selected_sheet]

    # Get the column index for student name and school name
    charstr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    while True:
        student_column = input("Which column is the student name in: ").upper()
        if student_column in charstr:
            break
        else:
            print("Please enter a valid alphabet.")
            continue

    while True:
        school_column = input("Which column is the school in: ").upper()
        if school_column in charstr:
            break
        else:
            print("Please enter a valid alphabet.")
            continue

    student_column_index = ord(student_column) - 65
    school_column_index = ord(school_column) - 65

    # Get first free column and row
    first_free_column = particulars_sheet.max_column + 1
    first_free_row = particulars_sheet.max_row + 1

    # Read cells from Excel, starting from the second row
    details = []
    for row in range(2, first_free_row):
        row_data = [
            particulars_sheet.cell(row=row, column=col).value
            for col in range(2, first_free_column)
        ]
        details.append(row_data)

    # Close Excel
    excel_instance.close()

    # Rest of your script remains unchanged

    # Initialize variables
    page_count = 1
    status = "Ok"

    # Loop while status is "Ok"
    while status == "Ok":
        while True:
            output_pdf_directory = input(
                "Enter the path of the directory you want to save split PDFs in: "
            ).strip('"')

            # Check if the filename is empty
            if output_pdf_directory == "":
                print("Filename cannot be empty.")
                continue

            # Check if the directory path exists
            if not os.path.exists(output_pdf_directory):
                print("Path does not exist")
                continue

            break

        for current_item in details:
            # Extract pages from PDF
            student_name = current_item[student_column_index - 1]
            school_name = current_item[school_column_index - 1]
            output_pdf_path = os.path.join(
                output_pdf_directory, f"{student_name}_{school_name}.pdf"
            )

            with open(selected_pdf_file, "rb") as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                pdf_writer = PyPDF2.PdfWriter()

                try:
                    pdf_writer.add_page(pdf_reader.pages[page_count - 1])
                except IndexError:
                    status = "OutOfBounds"
                    break

                # Save the extracted PDF
                with open(output_pdf_path, "wb") as output_pdf_file:
                    pdf_writer.write(output_pdf_file)

            page_count += 1
            status = "Completed"

    print("Process completed.")
except KeyboardInterrupt:
    print("\nProgram interrupted by user.")
